from string import ascii_uppercase

import numpy as np
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from setuptools import glob

# Headers for RAW_CSV files
HEADERS = ["Team", "Match_Num", "Auto_Cross", "Auto_Outer",
           "Auto_Bottom", "Tele_Outer", "Tele_Bottom", "Level",
           "Driver_Performance", "Auto_Perf", "Name", "Comments"]

# Excel fill colors using RGB values
darkRedFill = PatternFill(start_color='FF3333', end_color='FF3333', fill_type='solid')
redFill = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')
orangeFill = PatternFill(start_color='FFB266', end_color='FFB266', fill_type='solid')
yellowFill = PatternFill(start_color='FFFF66', end_color='FFFF66', fill_type='solid')
limeFill = PatternFill(start_color='B2FF66', end_color='B2FF66', fill_type='solid')
greenFill = PatternFill(start_color='66FF66', end_color='66FF66', fill_type='solid')


# Takes in a row from the DataFrame and calculates the total points using numpy
def total_points(pandaRow):
    CLIMB_POINTS = [0, 4, 6, 10, 15]
    Points = [2, 4, 2, 2, 1, CLIMB_POINTS[pandaRow.iloc[0][7]]]
    row = [col for col in pandaRow.iloc[0][2:8]]
    pandaRow["Total Points"] = sum(np.multiply(Points, row))
    return pandaRow


# Reads and concatenates RAW CSV files
def CSV_Reader():
    return pd.concat([total_points(pd.read_csv(file, names=HEADERS)) for file in glob.glob('New_CSVs/*.csv')],
                     ignore_index=True)


combinedData = CSV_Reader()
teams = combinedData.groupby("Team")

score_avg = teams[combinedData.columns[2:8]].mean()
total_avg = teams[combinedData.columns[12]].mean()


def teams_writer():
    teamData = dict(tuple(teams))
    with pd.ExcelWriter('Excel_Sheets/Teams.xlsx') as writer:
        for team in teamData.keys():
            team_average = pd.DataFrame(score_avg.loc[team]).swapaxes(0, 1)
            team_average["Total Points"] = [pd.DataFrame(total_avg).loc[team].get(key="Total Points")]
            teamData[team] = pd.concat([teamData[team], team_average])
            teamData[team].to_excel(writer, sheet_name="Team" + str(team), index=False)


def rankings_writer():
    total_avg.rename("Average Points", inplace=True)
    rankings = pd.concat([total_avg, score_avg], axis=1)
    rankings.sort_values(ascending=False, inplace=True, by="Average Points")

    writer = pd.ExcelWriter("Excel_Sheets/Teams.xlsx", engine='openpyxl')
    writer.book = load_workbook("Excel_Sheets/Teams.xlsx")
    rankings.to_excel(writer, sheet_name="Rankings")
    writer.save()

    workbook = openpyxl.load_workbook("Excel_Sheets/Teams.xlsx")
    ws = workbook["Rankings"]
    for column in ascii_uppercase[1:8]:
        column_mean = rankings[ws[column+'1'].value].mean()
        column_deviation = rankings[ws[column+'1'].value].std()
        rows = len(rankings.index) + 1
        ws.conditional_formatting.add("{L}2:{L}{rows}".format(L=column, rows=rows),
                                      CellIsRule(operator='lessThanOrEqual', formula=[column_mean - (3 * column_deviation)],
                                                 stopIfTrue=True, fill=darkRedFill))
        ws.conditional_formatting.add("{L}2:{L}{rows}".format(L=column, rows=rows),
                                      CellIsRule(operator='between', formula=[column_mean - (2 * column_deviation), column_mean - column_deviation],
                                                 stopIfTrue=True, fill=redFill))
        ws.conditional_formatting.add("{L}2:{L}{rows}".format(L=column, rows=rows),
                                      CellIsRule(operator='between', formula=[column_mean - column_deviation, column_mean],
                                                 stopIfTrue=True, fill=orangeFill))
        ws.conditional_formatting.add("{L}2:{L}{rows}".format(L=column, rows=rows),
                                      CellIsRule(operator='between', formula=[column_mean, column_mean + column_deviation],
                                                 stopIfTrue=True, fill=yellowFill))
        ws.conditional_formatting.add("{L}2:{L}{rows}".format(L=column, rows=rows),
                                      CellIsRule(operator='between', formula=[column_mean + column_deviation, column_mean + (2 * column_deviation)],
                                                 stopIfTrue=True, fill=limeFill))
        ws.conditional_formatting.add("{L}2:{L}{rows}".format(L=column, rows=rows),
                                      CellIsRule(operator='greaterThanOrEqual', formula=[column_mean + (2 * column_deviation)],
                                                 stopIfTrue=True, fill=greenFill))

    workbook.save("Excel_Sheets/Teams.xlsx")


def main():
    teams_writer()
    rankings_writer()


if __name__ == '__main__':
    main()
