import csv
from lib2to3.pgen2 import driver
import os
import openpyxl
from openpyxl import Workbook

# indexes in row
TEAM = 0
MATCH_NUM = 1
TARMAC = 2
AUTO_OUTTER = 3
AUTO_BOTTOM = 4
TELEOP_OUTTER = 5
TELEOP_BOTTOM = 6
LEVEL = 7
DRIVER_PERF = 8
AUTO_PERF = 9

CLIMB_POINTS = [0, 4, 6, 10, 15]
HEADER = ["Team", "Match_Num", "Auto_Cross", "Auto_Outter", "Auto_Bottom", "Tele_Outter", "Tele_Bottom", "Level", "Drvr_Perf", "Auto_Perf", "Name", "Comments", "Point_Contrib"]

team_points = {} # {"team number": [points]}
team_stats = {} # {"team number": [[round stat]]}

wb = Workbook() 

def read_csv(path):
    """Read the csv return the [team number, total point]"""
    file = open(path)
    reader = csv.reader(file)

    # header = next(reader)
    # header.append("Points Contrib")
    total_points = 0

    for row in reader:
        total_points += int(row[TARMAC]) * 2
        total_points += int(row[AUTO_OUTTER]) * 4
        total_points += int(row[AUTO_BOTTOM]) * 2
        total_points += int(row[TELEOP_BOTTOM]) * 2
        total_points += int(row[TELEOP_BOTTOM]) * 1
        total_points += CLIMB_POINTS[int(row[LEVEL])]

        row.append(total_points)
        if row[TEAM] not in team_stats.keys():
            team_stats[row[TEAM]] = [HEADER, row]
        else:
            team_stats[row[TEAM]].append(row)

        # return [row[TEAM], total_points]

def main():

    if os.path.exists("main.xlsx"):
        os.remove("main.xlsx")

    files = os.listdir("csv") # input folder
    for filename in files:
        read_csv("csv/" + filename)    
    write_excel(team_stats)
    

def write_excel(data: dict):
    """
        input: {team num: [[round 1 stat, ], [round 2 stat], ... ], team num: [,,,,,]]
    """
    row = 0
    col = 0
    m_row = 2
    master = wb.worksheets[0]
    
    master.cell(row=1, column=1).value = "Team #"
    master.cell(row=1, column=2).value = "Average"

    team_score = {}

    for team, stats in data.items():
        scores = []
        ws = wb.create_sheet("Team " + team)
        
        for d in stats:
            for a in d:
                # print(a)
                ws.cell(row=row+1, column=col+1).value = a
                col += 1
            if d[col-1] != "Point_Contrib":
                scores.append(int(d[col-1])) # added points to the scores array
            row += 1
            col = 0
        
        
        # ws.cell(row=row+1, column=12).value = "Average"
        ws.cell(row=row+1, column=13).value = sum(scores) / len(scores)
        ws.cell(row=row+1, column=1).value = "Average"
        col = 3
        for avg in generate_average(team_stats[team]):
            ws.cell(row=row+1, column=col).value = avg # fix alignment in excel
            col += 1    

        # write to the master sheet
        team_score[team] = sum(scores) / len(scores)

        row = 0
        col = 0

    team_score = sort_dict(team_score)
    print(team_score)
    for team, score in team_score.items():
        master.cell(row=m_row, column=1).value = team
        master.cell(row=m_row, column=2).value = score
        m_row += 1

    wb.save("main.xlsx")
    
def generate_average(stats):
    """
        stats: [[stat], [stat]]
        return [avg for each category]
    """
    tarmac = []
    auto_outter = []
    auto_bottom = []
    teleop_outter = []
    teleop_bottom = []
    level = []
    driver_perf = []
    auto_perf = []

    for stat in range(len(stats)):
        if stat == 0:
            continue
        tarmac.append(int(stats[stat][TARMAC]))
        auto_outter.append(int(stats[stat][AUTO_OUTTER]))
        auto_bottom.append(int(stats[stat][AUTO_BOTTOM]))
        teleop_outter.append(int(stats[stat][TELEOP_OUTTER]))
        teleop_bottom.append(int(stats[stat][TELEOP_BOTTOM]))
        level.append(int(stats[stat][LEVEL]))
        driver_perf.append(int(stats[stat][DRIVER_PERF]))
        auto_perf.append(int(stats[stat][AUTO_PERF]))

    return [
        average(tarmac),
        average(auto_outter),
        average(auto_bottom),
        average(teleop_outter),
        average(teleop_bottom),
        average(level),
        average(driver_perf),
        average(auto_perf)
    ]

def average(data):
    return sum(data)/len(data)

def sort_dict(dict1: dict):
    sorted_values = sorted(dict1.values())[::-1] # Sort the values
    sorted_dict = {}

    for i in sorted_values:
        for k in dict1.keys():
            if dict1[k] == i:
                sorted_dict[k] = dict1[k]
                break    
    
    return sorted_dict

if __name__ == "__main__":
    main()